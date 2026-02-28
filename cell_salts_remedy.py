import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cell Salt Remedies"

# Add headers
headers = ["Ailment", "Cell Salt Remedy", "Chemical Name", "Dosage", "Description"]
ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add remedy data
remedies = [
    ["Headache", "Potassium Phosphate", "K₃PO₄", "3-4 tablets every 2-3 hours", "Effective for mental fatigue and nervous headaches"],
    ["Fever", "Ferrum Phosphate", "Fe₃(PO₄)₂", "3-4 tablets every 1-2 hours", "First remedy in acute inflammation and fever"],
    ["Constipation", "Potassium Sulfate", "K₂SO₄", "3-4 tablets every 2 hours", "Aids in elimination and liver function"],
    ["Diarrhea", "Potassium Chloride", "KCl", "3-4 tablets every 1-2 hours", "Restores proper water balance in intestines"],
    ["Insomnia", "Magnesium Phosphate", "Mg₃(PO₄)₂", "3-4 tablets before bed", "Calms nervous system and promotes sleep"],
    ["Joint Pain", "Calcium Sulfate", "CaSO₄", "3-4 tablets 3 times daily", "Promotes healing of connective tissues"],
    ["Muscle Cramps", "Magnesium Phosphate", "Mg₃(PO₄)₂", "3-4 tablets every 30 minutes", "Relieves spasms and cramping"],
    ["Acne", "Silica", "SiO₂", "3-4 tablets 3 times daily", "Cleanses and purifies skin"],
    ["Cough", "Potassium Chloride", "KCl", "3-4 tablets every 1-2 hours", "Aids in bronchial and respiratory issues"],
    ["Anxiety", "Potassium Phosphate", "K₃PO₄", "3-4 tablets 3 times daily", "Balances nervous system"],
    ["Weakness", "Iron Phosphate", "Fe₃(PO₄)₂", "3-4 tablets 3 times daily", "Restores energy and vitality"],
    ["Inflammation", "Ferrum Phosphate", "Fe₃(PO₄)₂", "3-4 tablets every 2 hours", "First remedy for any inflammation"],
    ["Indigestion", "Potassium Chloride", "KCl", "3-4 tablets after meals", "Improves digestion and absorption"],
    ["Hair Loss", "Silica", "SiO₂", "3-4 tablets daily", "Strengthens hair and nails"],
    ["Brittle Nails", "Silica", "SiO₂", "3-4 tablets daily", "Promotes keratin and structural integrity"],
]

for remedy in remedies:
    ws.append(remedy)

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 35

# Save the workbook
wb.save("cell_salts_db.xlsx")
print("Excel file 'cell_salts.db' created successfully!")